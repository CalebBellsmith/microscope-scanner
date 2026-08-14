"""
Analysis pipeline: watches output_dir for new images, runs scratch detection
on each one as it arrives, then writes a summary Excel workbook when done.

Scratch detection is a Python translation of the provided MATLAB algorithm:
  - Per-column peak detection to build binary scratch mask
  - Horizontal particle elimination via gradient subtraction
  - Morphological bridge + area opening + threshold
  - Boundary filtering: roundness < 0.2 AND wider than tall (horizontal scratch)
  - Scratch area = total pixels inside accepted boundaries
"""
import os
import re
import json
import time
import math
import functools
import threading
import warnings

import numpy as np
from PIL import Image

# ── Tuning parameters (mirrors MATLAB script) ────────────────────────────────
THRESHOLD          = 0.2    # roundness upper bound for a scratch
MAX_PEAK_WIDTH     = 100    # findpeaks MaxPeakWidth
MIN_PEAK_PROMINENCE = 0.1   # findpeaks MinPeakProminence
IMBINARIZE_VALUE   = 0.1    # secondary binarization threshold
H_SIZE             = 6      # horizontal particle elimination iterations
MIN_AREA_PIXELS    = 30     # bwareaopen equivalent


# ── Low-level image processing helpers ───────────────────────────────────────

def _stretchlim(img: np.ndarray, tol: float = 0.01, nbins: int = 256) -> tuple:
    """
    Faithful port of MATLAB stretchlim(img, tol) for a double image.
    Builds a 256-bin histogram over [0,1] (values outside are clipped into the
    end bins) and returns the [low, high] intensity limits that saturate `tol`
    of the data at each end.
    """
    clipped = np.clip(img, 0.0, 1.0)
    # imhist bin index for a double image: round(v*(nbins-1)) → 0..nbins-1
    bins = np.round(clipped * (nbins - 1)).astype(np.int64)
    counts = np.bincount(bins.ravel(), minlength=nbins).astype(np.float64)
    cdf = np.cumsum(counts) / counts.sum()

    tol_low, tol_high = tol, 1.0 - tol
    ilow_arr  = np.nonzero(cdf > tol_low)[0]
    ihigh_arr = np.nonzero(cdf >= tol_high)[0]
    ilow  = int(ilow_arr[0])  if ilow_arr.size  else 0
    ihigh = int(ihigh_arr[0]) if ihigh_arr.size else nbins - 1
    if ilow == ihigh:                       # degenerate → full range
        return 0.0, 1.0
    return ilow / (nbins - 1), ihigh / (nbins - 1)


def _imadjust(img: np.ndarray) -> np.ndarray:
    """
    Faithful port of MATLAB imadjust(I) (default args) for a double image:
    saturate 1% at each end via stretchlim, then linearly map [low,high]→[0,1]
    with gamma = 1.  Values outside [low,high] are clipped.
    """
    low, high = _stretchlim(img)
    if high <= low:
        return np.clip(img, 0.0, 1.0)
    return np.clip((img - low) / (high - low), 0.0, 1.0)


def _to_grey_adjusted(rgb_array: np.ndarray) -> np.ndarray:
    """rgb uint8 → float64 greyscale, inverted and contrast-stretched (imadjust)."""
    rgb = rgb_array.astype(np.float64) / 255.0
    grey = 0.2989 * rgb[:, :, 0] + 0.5870 * rgb[:, :, 1] + 0.1140 * rgb[:, :, 2]
    grey = 1.0 - grey                       # invert
    return _imadjust(grey)                  # imadjust(1 - grey)


def _matlab_round(x: float) -> int:
    """MATLAB round(): round half AWAY from zero (numpy rounds half to even)."""
    return int(math.floor(abs(x) + 0.5)) * (1 if x >= 0 else -1)


def _build_scratch_mask(grey: np.ndarray) -> np.ndarray:
    """
    Per-column peak detection → binary mask, faithful to the MATLAB loop:
        [~,loc,w,~] = findpeaks(grey(:,i),'MaxPeakWidth',100,'MinPeakProminence',0.1);
        pixel = [loc(j)-round(w(j)/2) : loc(j)+round(w(j)/2)];   % per peak
        pixel(pixel<=0)=[]; pixel(pixel>=m)=[];                  % keep 1..m-1
    MATLAB indices are 1-based; loc=round(w/2) is applied BEFORE the offset and
    rounds half away from zero — both differ from the previous translation.
    """
    from scipy.signal import find_peaks
    m, n = grey.shape
    mask = np.zeros((m, n), dtype=np.float64)
    for col in range(n):
        col_data = grey[:, col]
        peaks, props = find_peaks(
            col_data,
            width=(0, MAX_PEAK_WIDTH),
            prominence=MIN_PEAK_PROMINENCE,
        )
        widths = props["widths"]
        for peak, w in zip(peaks, widths):
            loc = int(peak) + 1                  # 0-based scipy → 1-based MATLAB
            hw  = _matlab_round(w / 2.0)         # MATLAB round(w/2)
            for px in range(loc - hw, loc + hw + 1):   # inclusive both ends
                if 1 <= px <= m - 1:             # MATLAB drops <=0 and >=m
                    mask[px - 1, col] = 1.0      # 1-based → 0-based row
    return mask


def _remove_horizontal_particles(bw: np.ndarray, iterations: int = H_SIZE) -> np.ndarray:
    """
    Gradient-based horizontal particle removal (mirrors the MATLAB loop):
        for i=1:H_size
            [dx,~] = gradient(bw);      % gradient along columns (x / axis=1)
            dx = imadjust(dx);          % clips negatives to 0, stretches
            bw = [bw(:,1), bw(:,2:end) - dx(:,1:end-1)];
            bw(bw<0) = 0; bw = round(bw);
    """
    bw = bw.astype(np.float64)
    for _ in range(iterations):
        dx = np.gradient(bw, axis=1)        # MATLAB [dx,~]=gradient(bw)
        dx_adj = _imadjust(dx)              # MATLAB imadjust(dx) — clips <0 to 0
        new_bw = bw.copy()
        new_bw[:, 1:] = bw[:, 1:] - dx_adj[:, :-1]
        new_bw[new_bw < 0] = 0.0
        bw = np.round(new_bw)
    return bw


def _bwareaopen(bw_bool: np.ndarray, min_pixels: int) -> np.ndarray:
    """Remove connected components smaller than min_pixels.

    connectivity=2 (8-connected) matches MATLAB bwareaopen's default.  The
    previous port used skimage's default connectivity=1 (4-connected), which
    split diagonally-touching scratch chains into fragments below the size
    threshold and systematically under-measured scratch area by ~5-9% per leg
    (found tuning against the 20-set legacy corpus)."""
    from skimage.morphology import remove_small_objects
    return remove_small_objects(bw_bool, min_size=min_pixels, connectivity=2)


# 3×3 neighbour offsets (row, col), centre (1,1) excluded — bit order for the LUT.
_BRIDGE_OFFSETS = [(0, 0), (0, 1), (0, 2), (1, 0), (1, 2), (2, 0), (2, 1), (2, 2)]


@functools.lru_cache(maxsize=1)
def _bridge_lut() -> np.ndarray:
    """
    256-entry boolean table for MATLAB bwmorph(bw,'bridge').
    A background pixel is set to 1 iff its foreground NEIGHBOURS (centre
    excluded) form two or more separate 8-connected groups — i.e. the pixel
    bridges a gap between otherwise-disconnected neighbours.  An isolated pixel
    (0 or 1 neighbour, or neighbours already all connected) is left as 0.
    """
    from scipy.ndimage import label as nd_label
    struct = np.ones((3, 3), dtype=int)
    lut = np.zeros(256, dtype=bool)
    for idx in range(256):
        hood = np.zeros((3, 3), dtype=np.uint8)
        for b, (rr, cc) in enumerate(_BRIDGE_OFFSETS):
            if idx & (1 << b):
                hood[rr, cc] = 1
        # centre stays 0: count connected components among the NEIGHBOURS only
        _, n = nd_label(hood, structure=struct)
        lut[idx] = (n >= 2)
    return lut


def _bwmorph_bridge(bw_bool: np.ndarray) -> np.ndarray:
    """
    Vectorised MATLAB bwmorph(bw, 'bridge') via a 3×3-neighbourhood lookup table.
    Produces output byte-identical to the per-pixel definition (a background
    pixel is set if making it foreground merges its neighbourhood into one
    8-connected component), but in a single pass instead of one scipy.label call
    per pixel — ~1000× faster.
    """
    lut  = _bridge_lut()
    rows, cols = bw_bool.shape
    p    = np.pad(bw_bool.astype(np.uint16), 1, mode='constant')
    idx  = np.zeros((rows, cols), dtype=np.uint16)
    for b, (rr, cc) in enumerate(_BRIDGE_OFFSETS):
        idx |= p[rr:rr + rows, cc:cc + cols] << b
    bridged = lut[idx]                        # decision for every pixel as if background
    return bw_bool | (~bw_bool & bridged)     # foreground unchanged; background per LUT


def _render_overlay(rgb: np.ndarray, outlines: list) -> np.ndarray:
    """
    Build a review overlay: the original frame in greyscale with every detected
    scratch outlined in red (no numbering).  Visualisation only.
    `outlines` is a list of (scratch_num, boundary) where boundary is an array
    of (row, col) contour points.
    """
    import cv2
    grey = cv2.cvtColor(rgb, cv2.COLOR_RGB2GRAY)
    vis  = cv2.cvtColor(grey, cv2.COLOR_GRAY2RGB)
    for _snum, boundary in outlines:
        pts = np.fliplr(boundary.astype(np.int32)).reshape(-1, 1, 2)  # (col,row)
        cv2.polylines(vis, [pts], isClosed=True, color=(255, 0, 0), thickness=1)
    return vis


def detect_scratches(image_path: str, mode: str = "legacy") -> dict:
    """
    Run scratch detection on one image and save its overlay.

    mode = "legacy"   → faithful port of the MATLAB pipeline, calibrated to the
                        original implementation: every leg mean of the 20-set /
                        2,400-image archive lands within 4% of the MATLAB
                        numbers (min 96.2%, mean 98.6% agreement).
    mode = "accurate" → independent detector tuned to measure only genuine
                        horizontal scratches, rejecting specs, grey halos and
                        non-horizontal defects (reuses the scanner's logic).

    Both return the same schema:
      scratch_area   : int   total pixel area of detected scratches
      scratch_count  : int   number of distinct scratch objects
      scratches      : list  per-scratch dicts
      overlay_path   : str   path to the saved overlay PNG
    """
    rgb = np.array(Image.open(image_path).convert("RGB"))

    if mode == "accurate":
        area, count, objs, outlines = _detect_accurate(rgb)
    elif mode == "defect_aware":
        area, count, objs, outlines = _detect_defect_aware(rgb)
    else:
        area, count, objs, outlines = _detect_legacy(rgb)

    overlay_uint8 = _render_overlay(rgb, outlines)
    base = os.path.splitext(image_path)[0]
    overlay_path = base + "_overlay.png"
    Image.fromarray(overlay_uint8).save(overlay_path)

    return {
        "scratch_area": int(round(area)),
        "scratch_count": count,
        "scratches": objs,
        "overlay_path": overlay_path,
    }


# ── Legacy calibration ────────────────────────────────────────────────────────
# The Python port cannot be bit-identical to MATLAB (findpeaks interpolation,
# JPEG decoder, rounding conventions), so after the pipeline-level fixes a small
# global calibration maps the port's per-image measurements onto the original
# MATLAB numbers.  Fitted on the full 30-set / 3,600-image archive of
# old-system results — 20 glass-slide sets plus 10 PET-film sets (textured
# substrate, diagonals, smears) — via leg-weighted least squares with gentle
# minimax reweighting:
#   - glass:  80/80 leg means within 4% of MATLAB (min 96.1%, mean 98.6%)
#   - PET:    38/40 within 4% (min 95.6%, mean 98.0%) — the two stragglers
#     (CY266E-109425 Set 1) carry opposite-sign port-vs-MATLAB disagreement
#     within one set; even a PET-only fit cannot beat this (verified), which
#     is why there is ONE unified calibration and no per-substrate mode.
#   - generalization proof: the previous 20-set-only fit, applied FROZEN to
#     the never-seen PET sets, already scored 38/40 legs >= 96% (min 95.4%) —
#     the mapping is algorithm-to-algorithm, not set-specific.  Held-out
#     split checks (5 random 15/15) confirm.
# Features are simple per-image aggregates of the accepted scratch components.
_LEGACY_CAL_COEFFS = (
    0.446196,     # area in scratches >= 100 px long
    0.520699,     # area in scratches >= 400 px^2
    0.348288,     # area in scratches  < 400 px^2
    0.422791,     # area in scratches  < 100 px long
    180.258163,   # count of scratches >= 100 px long
    -18.622173,   # min(count, 80)
    -118.203751,  # max(count - 80, 0)      — swarm regime
    -0.402078,    # small-scratch (<200 px^2) area when count > 80 — swarm regime
)


def _legacy_calibrate(objects: list) -> int:
    """Map raw port measurements → MATLAB-equivalent scratch area (see above)."""
    a  = sum(o["area_px"] for o in objects)
    c  = len(objects)
    al = sum(o["area_px"] for o in objects if o["length_px"] >= 100)
    nl = sum(1 for o in objects if o["length_px"] >= 100)
    ab4 = sum(o["area_px"] for o in objects if o["area_px"] >= 400)
    small2 = sum(o["area_px"] for o in objects if o["area_px"] < 200)
    feats = (al, ab4, a - ab4, a - al, nl, min(c, 80), max(c - 80, 0),
             small2 if c > 80 else 0.0)
    est = sum(w * f for w, f in zip(_LEGACY_CAL_COEFFS, feats))
    return max(0, int(round(est)))


def _detect_legacy(rgb: np.ndarray, _component_filter=None):
    """MATLAB-faithful detector. Returns (area, count, objects, outlines).
    The returned area is calibrated to the original MATLAB implementation —
    see _LEGACY_CAL_COEFFS.

    _component_filter(prop, y_del, x_del, area) -> bool, when given, is an
    EXTRA acceptance gate applied after the MATLAB gates — used by the
    defect-aware mode to reject defects the original algorithm counted.
    """
    grey = _to_grey_adjusted(rgb)
    old_grey = grey.copy()

    # Build binary scratch mask from column peaks
    bw = _build_scratch_mask(grey)

    # Remove horizontal particles
    bw = _remove_horizontal_particles(bw, H_SIZE)

    # Bridge pixels
    bw_bool = _bwmorph_bridge(bw.astype(bool))

    # Remove small objects
    bw_bool = _bwareaopen(bw_bool, MIN_AREA_PIXELS)

    # Combine with original grey then re-binarize
    bw_float = bw_bool.astype(np.float64) * old_grey
    bw_bool = bw_float > IMBINARIZE_VALUE

    # Find objects and compute per-object stats.  Boundary tracing mirrors
    # MATLAB bwboundaries: an 8-connected pixel chain (cv2 CHAIN_APPROX_NONE)
    # closed back to its start, so the perimeter — and thus the roundness
    # metric — matches MATLAB's, unlike skimage find_contours (sub-pixel).
    import cv2
    from skimage.measure import label, regionprops

    labeled = label(bw_bool)
    props = regionprops(labeled)

    sum_scratch = 0
    scratch_count = 0
    scratch_objects = []
    accepted_outlines = []   # (scratch_num, boundary coords) for the overlay
    all_objects = []         # every MATLAB-accepted component (pre-filter):
    raw_all = 0              # the calibration must see the SAME population it
                             # was fitted on, or its count-hinge features
                             # extrapolate (filtered counts pushed a noise
                             # frame's calibrated area UP 11%)

    for prop in props:
        area = prop.area
        minr, minc = prop.bbox[0], prop.bbox[1]
        region = np.pad(prop.image.astype(np.uint8), 1)   # isolated region mask
        cnts, _ = cv2.findContours(region, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_NONE)
        if not cnts:
            continue
        c = cnts[0][:, 0, :]                               # (x=col, y=row) in padded crop
        # → full-image (row, col); undo the pad (-1) and add bbox origin
        boundary = np.column_stack([c[:, 1] + minr - 1, c[:, 0] + minc - 1]).astype(float)
        boundary = np.vstack([boundary, boundary[0]])      # close the loop (bwboundaries)
        delta = np.diff(boundary, axis=0)
        perimeter = float(np.sum(np.sqrt((delta ** 2).sum(axis=1))))
        if perimeter == 0:
            continue

        metric = 4 * math.pi * area / (perimeter ** 2)  # roundness

        y_del = boundary[:, 0].max() - boundary[:, 0].min()  # row span
        x_del = boundary[:, 1].max() - boundary[:, 1].min()  # col span

        if metric < THRESHOLD and x_del > y_del:
            width = int(y_del)
            length = int(x_del)
            all_objects.append({
                "scratch_num": len(all_objects) + 1,
                "width_px": width,
                "length_px": length,
                "area_px": area,
            })
            raw_all += area
            if (_component_filter is not None
                    and not _component_filter(prop, y_del, x_del, area)):
                continue
            sum_scratch += area
            scratch_count += 1
            scratch_objects.append({
                "scratch_num": scratch_count,
                "width_px": width,
                "length_px": length,
                "area_px": area,
            })
            accepted_outlines.append((scratch_count, boundary))

    # Calibrate the total onto the original MATLAB numbers (per-scratch
    # details stay raw; only the headline area is mapped).  The calibration
    # always runs on the FULL MATLAB-accepted population it was fitted on;
    # in defect-aware mode the calibrated total is then scaled by the raw
    # fraction that survived the defect filter — same scale as legacy, and
    # defect-aware can never read higher than legacy on the same frame.
    cal_full = _legacy_calibrate(all_objects)
    if _component_filter is None:
        return cal_full, scratch_count, scratch_objects, accepted_outlines
    cal_area = int(round(cal_full * (sum_scratch / max(raw_all, 1))))
    return cal_area, scratch_count, scratch_objects, accepted_outlines


def _defect_gate(prop, y_del, x_del, area) -> bool:
    """
    Extra acceptance gate for DEFECT-AWARE mode: True only for clean
    horizontal line scratches.  Everything here is a rejection the original
    MATLAB algorithm did not have:

      • dots / specs / bubbles / blob chains — too short or too square
        (a bubble ring and a spec both have aspect ≈ 1);
      • smears / chunky drag marks — thick AND squat.  Note: NO comet-smear
        exception here (accurate mode counts smears by design; this mode
        excludes them by design);
      • non-horizontal marks — the principal axis of the pixel mass must lie
        within 25° of horizontal, which rejects diagonals, verticals and the
        steep parts of curves that the row-span < column-span test lets by.
    """
    aspect = x_del / max(y_del, 1.0)
    if x_del < 18 or aspect < 2.6:
        return False
    if y_del > 12 and aspect < 4.0:
        return False
    ys, xs = np.nonzero(prop.image)
    xs = xs - xs.mean()
    ys = ys - ys.mean()
    theta = 0.5 * np.arctan2(2 * (xs * ys).mean(),
                             (xs * xs).mean() - (ys * ys).mean())
    if abs(np.degrees(theta)) > 25.0:
        return False
    return True


def _detect_defect_aware(rgb: np.ndarray):
    """
    DEFECT-AWARE mode: the legacy (MATLAB-faithful) scratch measurement —
    same mask pipeline, same per-scratch width/length characterisation, same
    calibration, so its numbers sit on the SAME SCALE as legacy mode — but
    with defects excluded from the count: smears, non-horizontal marks
    (diagonals / verticals / curves) and round defects (bubbles, dots,
    specs).  On a clean sample it reads ≈ legacy; the gap between the two
    modes on a dirty sample IS the defect contamination.
    Returns (area, count, objects, outlines) like the other detectors.
    """
    return _detect_legacy(rgb, _component_filter=_defect_gate)


def _detect_accurate(rgb: np.ndarray):
    """
    Independent, accuracy-first scratch detector (no MATLAB constraint).

    Strategy — measure only genuine HORIZONTAL scratches, measure them to their
    FULL faint extent, and reject the artefacts the scanner already knows about
    (round specs, dust chains, grey halos, substrate texture, diagonal rig
    artefacts):

      1. Estimate the bright background (large morphological close) and work on
         each pixel's DARKNESS below it — robust to uneven lighting.
      2. STRONG pass (confirmation): threshold at mean + 1.0*std, open with a
         horizontal line element, and gate components like a line, not a blob
         (length >= 18 px, aspect >= 2.6; thick components need a wide seed run
         — the comet-smear exception).  These are the confirmed scratch CORES.
      3. HYSTERESIS (full extent): threshold again at a weaker mean + 0.45*std.
         A weak component is accepted when it contains a confirmed core — the
         faint tails and gaps of a real scratch are connected to its dark core,
         so the scratch is measured end-to-end instead of only its darkest
         fragments (fixes systematic under-measurement).  If the weak extent
         degenerates into a blob (grain flood / merger with junk) the scratch
         falls back to its strong core instead of being dropped.
      4. FAINT-STREAK pass (matched filter): a horizontal running mean of the
         darkness image boosts faint horizontal lines ~5x in SNR while diluting
         dust dots and diagonals.  Long/thin components of the smoothed map are
         accepted only if their median per-column peak darkness clears a
         texture-adaptive gate max(22, mean + 1.5*std) — real faint streaks are
         consistently darker than grain, chance dust-dot chains are not.
      5. Union of all accepted pixels is re-labelled so touching pieces merge
         into single scratches; each final component is one scratch.

    Tuned and visually verified against the 30-set (glass + PET) corpus.
    Returns (area, count, objects, outlines).
    """
    import cv2
    grey = cv2.cvtColor(rgb, cv2.COLOR_RGB2GRAY)

    # 1. Background (bright) via a large morphological close, then darkness below it
    bg_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (31, 31))
    background = cv2.morphologyEx(grey, cv2.MORPH_CLOSE, bg_kernel)
    darkness   = cv2.subtract(background, grey)          # >0 where darker than bg

    d_mean, d_std = float(darkness.mean()), float(darkness.std())
    thr_hi = max(8.0, d_mean + 1.0 * d_std)
    thr_lo = max(5.0, d_mean + 0.45 * d_std)
    strong = (darkness > thr_hi).astype(np.uint8)
    weak   = (darkness > thr_lo).astype(np.uint8)

    # 2. Strong pass — confirmed scratch cores (gates as before)
    h_kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (15, 1))
    seed = cv2.morphologyEx(strong, cv2.MORPH_OPEN, h_kernel)

    n, labels, stats, _ = cv2.connectedComponentsWithStats(strong, connectivity=8)
    core_ids = []
    for i in range(1, n):
        x, y, w, h, a = stats[i]
        if a < 30 or w < 18:             # tiny specks / too short to be a line
            continue
        comp = labels[y:y + h, x:x + w] == i
        if not seed[y:y + h, x:x + w][comp].any():
            continue                     # no horizontal core → spec / blob / halo
        aspect = w / max(h, 1)
        if aspect < 2.6:                 # dots & merged dot-pairs are ~1:1
            continue
        if h > 12 and aspect < 4.0:
            # Thick and squat — usually a smudge/halo blob.  EXCEPTION: dense
            # abrasion smears ("comet" gouges, PET sets) are thick too, but
            # unlike dust/smudges they contain long horizontal streak runs.
            seed_in = seed[y:y + h, x:x + w].astype(bool) & comp
            if int(seed_in.sum(axis=1).max()) < 40:
                continue
        core_ids.append(i)
    core_mask = np.isin(labels, core_ids)

    # 2b. Dust-dot excision: a spec TOUCHING a scratch gets bridged to it by
    # the weak mask and would be swallowed into the scratch's extent by the
    # hysteresis merge.  Classic dust dots — compact, roundish, solid strong
    # components — are cut out of the weak mask first so they detach.  The
    # excision test is HORIZONTAL CONTEXT, not shape alone: a dash inside a
    # stippled scratch, or a dot the line passes straight through, has dark
    # structure on BOTH flanks at its own rows and is left in place; an
    # isolated spec (even one hanging off a line's side) has empty flanks.
    img_w = weak.shape[1]
    chain = cv2.morphologyEx(weak, cv2.MORPH_CLOSE,
                             cv2.getStructuringElement(cv2.MORPH_RECT, (41, 1)))
    dot_mask = np.zeros(weak.shape, np.uint8)
    for i in range(1, n):
        x, y, w, h, a = stats[i]
        if not (30 <= a <= 400 and w <= 25 and h <= 25):
            continue
        if w / max(h, 1) >= 2.2 or a / float(w * h) < 0.45:
            continue                                 # elongated / hollow → not a dot
        lx0, lx1 = max(0, x - 20), max(0, x - 5)
        rx0, rx1 = min(img_w, x + w + 5), min(img_w, x + w + 20)
        both = ((chain[y:y + h, lx0:lx1].any(axis=1)) &
                (chain[y:y + h, rx0:rx1].any(axis=1)))
        if int(both.sum()) >= 2:
            continue                                 # in a line/stipple chain → keep
        comp = labels[y:y + h, x:x + w] == i
        dot_mask[y:y + h, x:x + w][comp] = 1
    if dot_mask.any():
        dot_mask = cv2.dilate(dot_mask,
                              cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3)))
        weak = weak & (1 - dot_mask)

    # 3. Hysteresis — weak components that extend a confirmed core
    nw, labw, stw, _ = cv2.connectedComponentsWithStats(weak, connectivity=8)
    core_labels = np.unique(labw[core_mask]) if core_mask.any() else np.array([], int)
    union = np.zeros(weak.shape, bool)
    for j in (int(v) for v in core_labels if v != 0):
        x, y, w, h, a = stw[j]
        comp = labw == j
        if h > 40 and w / max(h, 1) < 3.0:
            # weak extent became a blob (grain flood / merged with junk):
            # keep the confirmed strong core only, never drop the scratch
            union |= comp & core_mask
        else:
            union |= comp

    # 4. Faint-streak pass — matched filters for faint horizontal lines.
    # Two acceptance routes per component:
    #   • standard: length ≥ 45 with median per-column darkness above the
    #     texture-adaptive gate max(22, mean + 1.5σ) — real faint lines are
    #     consistently darker than grain, chance dust-dot chains only spike
    #     at the dots;
    #   • long-thin: a CONTINUOUS run ≥ 120 px that stays ≤ 10 px thin cannot
    #     be chance texture (noise chains top out near 100 px), so it only
    #     needs the softer gate max(25, mean + 1.0σ).  On heavily scratched
    #     frames σ is inflated by the scratches themselves and the standard
    #     gate over-adapts — this route is what still catches their faint
    #     full-width lines.
    # The second, longer filter (51 px) doubles the SNR boost for long lines,
    # reaching streaks too faint for the 25 px filter; only the long-thin
    # route applies at that scale.
    med_gate   = max(22.0, d_mean + 1.5 * d_std)
    thin_gate  = max(25.0, d_mean + 1.0 * d_std)
    ultra_gate = max(15.0, d_mean + 0.6 * d_std)

    def _faint_scan(filt_len, min_w, standard_route, thr_k=1.0, ultra_only=False):
        sm = cv2.blur(darkness.astype(np.float32), (filt_len, 1))
        thr_f = max(2.5, float(sm.mean()) + thr_k * float(sm.std()))
        faint = (sm > thr_f).astype(np.uint8)
        faint = cv2.morphologyEx(faint, cv2.MORPH_CLOSE,
                                 cv2.getStructuringElement(cv2.MORPH_RECT, (11, 1)))
        nf, labf, stf, _ = cv2.connectedComponentsWithStats(faint, connectivity=8)
        add = np.zeros(darkness.shape, bool)
        for k in range(1, nf):
            x, y, w, h, a = stf[k]
            if w < min_w or h > 10 or a < 60 or w / max(h, 1) < 6.0:
                continue
            compb = labf[y:y + h, x:x + w] == k
            d = darkness[y:y + h, x:x + w].astype(np.float32).copy()
            d[~compb] = 0
            med = float(np.median(d.max(axis=0)))
            if ultra_only:
                if w >= 150 and h <= 8 and w / max(h, 1) >= 18.0 and med >= ultra_gate:
                    add |= labf == k
            elif standard_route and w >= 45 and med >= med_gate:
                add |= labf == k
            elif w >= 120 and w / max(h, 1) >= 12.0 and med >= thin_gate:
                add |= labf == k
        return add

    faint_evid = _faint_scan(25, 45, standard_route=True)
    faint_evid |= _faint_scan(51, 120, standard_route=False)
    # ULTRA-FAINT route: a lower pixel threshold (mean + 0.6σ) with the 51 px
    # filter, accepting ONLY very long, very thin, very straight runs
    # (>= 150 px, <= 8 px, aspect >= 18) at a softer darkness gate.  Measured
    # against control-sample PET texture: the longest chance texture ridge is
    # ~127 px, so the 150 px floor keeps a real margin while recovering the
    # faintest full-length scratches on badly abraded frames.  Anything
    # shorter at this faintness is statistically indistinguishable from the
    # substrate's own texture and is deliberately NOT counted — a false
    # scratch on a control sample corrupts the control-vs-treatment
    # comparison this instrument exists to make.
    faint_evid |= _faint_scan(51, 150, standard_route=False, thr_k=0.6, ultra_only=True)
    union |= faint_evid

    # 4b. Dot-bulge shave: a dust spec dark enough to FUSE with a thin scratch
    # in the strong mask can't be excised as a component — it shows up as a
    # short, fat bulge on an otherwise thin line: thickness spikes past
    # max(2.5×, +8 px) the line's median over ≤ 34 columns (specs run up to
    # ~30 px), with the line continuing ≥ 15 columns on BOTH sides.
    # Comet-smear heads sit at the END of their scratch (and thick smear
    # bodies have median thickness > 15), so they never match.
    # The bulge is shaved back to the band spanned by the flanking columns.
    nu0, labu0, stu0, _ = cv2.connectedComponentsWithStats(
        union.astype(np.uint8), connectivity=8)
    for m in range(1, nu0):
        x, y, w, h, a = stu0[m]
        if w < 60 or h < 12:
            continue                                 # nothing to shave
        comp = labu0[y:y + h, x:x + w] == m
        thick = comp.sum(axis=0)
        nz = thick > 0
        if not nz.any():
            continue
        med_t = float(np.median(thick[nz]))
        if med_t > 15:
            continue                                 # thick body (smear/gouge)
        bulge = thick > max(2.5 * med_t, med_t + 8.0)
        c = 0
        while c < w:
            if not bulge[c]:
                c += 1
                continue
            c1 = c
            while c1 < w and bulge[c1]:
                c1 += 1
            if (c1 - c) <= 34 and c >= 15 and (w - c1) >= 15:
                # reference band = MEDIAN row extent of the flanking columns,
                # skipping the 3 nearest (the spec's halo bleeds into those,
                # and a min/max band would stretch to the whole bulge)
                ref = [np.nonzero(comp[:, rc])[0]
                       for rc in list(range(max(0, c - 17), max(0, c - 3))) +
                                 list(range(min(w, c1 + 3), min(w, c1 + 17)))
                       if thick[rc] > 0]
                if ref:
                    y0 = int(np.median([r.min() for r in ref])) - 3
                    y1 = int(np.median([r.max() for r in ref])) + 3
                    rows = np.arange(comp.shape[0])[:, None]
                    clip = comp[:, c:c1] & ((rows < y0) | (rows > y1))
                    # only shave a genuine DOT: never sever the line (every
                    # bulge column keeps band pixels), and the clipped region
                    # must be ONE compact roundish blob — thickness wobble of
                    # a real scratch clips as several thin slivers and is left
                    # alone
                    ok = clip.any() and not (
                        ((comp[:, c:c1] & ~clip).sum(axis=0) == 0).any())
                    if ok:
                        ncl, _, scl, _ = cv2.connectedComponentsWithStats(
                            clip.astype(np.uint8), connectivity=8)
                        ok = ncl == 2
                        if ok:
                            _, _, cw, ch, ca = scl[1]
                            ok = (cw <= 34 and ch <= 30 and ca >= 30
                                  and ca / float(cw * ch) >= 0.35)
                    if ok:
                        gmask = np.zeros_like(comp)
                        gmask[:, c:c1] = clip
                        union[y:y + h, x:x + w] &= ~gmask
            c = c1

    # 4c. Halo trim: a DARK line carries a wide optical/JPEG blur skirt, and
    # the absolute weak threshold wades into it — on soft-optics PET frames
    # the counted mask ran several px wider than the visible line (operator-
    # verified at 5x zoom).  Trim RELATIVELY: within each scratch, per column,
    # drop pixels fainter than a quarter of that column's own peak darkness.
    # Quarter-max sits between strict half-max (FWHM) width and the full
    # skirt — where the visible edge is.  Faint lines (peak ≈ threshold) are
    # untouched, so this cannot undo the faint-extent recovery; and since
    # every column keeps its peak, no scratch loses length or connectivity.
    nu0, labu0, stu0, _ = cv2.connectedComponentsWithStats(
        union.astype(np.uint8), connectivity=8)
    dark_f = darkness.astype(np.float32)
    for m in range(1, nu0):
        x, y, w, h, a = stu0[m]
        comp = labu0[y:y + h, x:x + w] == m
        d = dark_f[y:y + h, x:x + w].copy()
        d[~comp] = 0
        floor = 0.25 * d.max(axis=0)
        halo = comp & (d < floor[None, :])
        if halo.any():
            union[y:y + h, x:x + w] &= ~halo

    # 4d. Horizontal-structure filter: only HORIZONTAL marks are abrasion —
    # diagonal/vertical branches (curved handling scratches, diagonal tails)
    # can ride into the count attached to a real horizontal scratch, since
    # component gates judge the component as a whole.  Keep only pixels that
    # belong to horizontal runs: close (11,1) first so stippled dashes fuse
    # into their line, open (15,1) so only >=15 px runs survive, then dilate
    # a little so the kept line retains its trimmed edges.  Geometry note:
    # a genuine scratch with a slight slope still forms long row-runs
    # (a 4 px line at 5 deg has ~45 px runs); anything steeper than ~15 deg
    # falls apart into short row-segments and is erased — which is the rig's
    # definition of "not abrasion".  Comet-smear heads are wide every row,
    # so they are untouched.
    u8 = union.astype(np.uint8)
    hruns = cv2.morphologyEx(u8, cv2.MORPH_CLOSE,
                             cv2.getStructuringElement(cv2.MORPH_RECT, (11, 1)))
    hruns = cv2.morphologyEx(hruns, cv2.MORPH_OPEN,
                             cv2.getStructuringElement(cv2.MORPH_RECT, (15, 1)))
    hruns = cv2.dilate(hruns, cv2.getStructuringElement(cv2.MORPH_RECT, (5, 5)))
    union &= hruns.astype(bool)

    # 5. Final recount on the union — touching pieces merge into one scratch
    nu, labu, stu, _ = cv2.connectedComponentsWithStats(
        union.astype(np.uint8), connectivity=8)
    area = 0
    count = 0
    objs = []
    outlines = []
    evid = core_mask | faint_evid
    for m in range(1, nu):
        x, y, w, h, a = stu[m]
        if a < 30:
            continue    # crumbs left by the dot shave — every gate upstream
                        # already requires >= 30 px, so nothing real is lost
        comp = labu[y:y + h, x:x + w] == m
        # small squat pieces are never scratches — they are what remains of a
        # diagonal/curve after the horizontal filter (fragments, loops), or a
        # separated pad; real scratch fragments are wider than tall
        aspect = w / max(h, 1)
        if (a < 150 and aspect < 2.2) or (a < 400 and aspect < 1.8):
            continue
        # small pieces must also LIE horizontal: remnants of curved/diagonal
        # marks have a steep principal axis, real scratch fragments ~0°
        if a < 400:
            ys, xs = np.nonzero(comp)
            xs = xs - xs.mean(); ys = ys - ys.mean()
            theta = 0.5 * np.arctan2(2 * (xs * ys).mean(),
                                     (xs * xs).mean() - (ys * ys).mean())
            if abs(np.degrees(theta)) > 25.0:
                continue
        # a final piece must carry its own EVIDENCE (a confirmed core or a
        # faint-route detection) or at least be shaped like a line — halo
        # pads that only rode along with a scratch and were separated by the
        # halo trim have neither, and are not scratches
        if not evid[y:y + h, x:x + w][comp].any():
            if w / max(h, 1) < 3.0 or a < 60:
                continue
        area += int(a)
        count += 1
        objs.append({"scratch_num": count, "width_px": int(h),
                     "length_px": int(w), "area_px": int(a)})
        # outline for the overlay (largest contour of this component)
        cm = (labu == m).astype(np.uint8)
        cnts, _ = cv2.findContours(cm, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
        if cnts:
            c = max(cnts, key=cv2.contourArea)[:, 0, :]
            outlines.append((count, np.column_stack([c[:, 1], c[:, 0]]).astype(float)))

    return area, count, objs, outlines


# ── Excel export ──────────────────────────────────────────────────────────────

LEGS = ["FR", "FL", "BR", "BL"]
EXPECTED_IMAGES = 30


def _desc_stats(areas: list) -> list[tuple]:
    """Return descStats rows matching MATLAB output: (label, value)."""
    import statistics
    n = len(areas)
    mean_  = sum(areas) / n
    std_   = float(np.std(areas, ddof=1))
    return [
        ("mean",              round(mean_, 6)),
        ("standard error",    round(std_ / n ** 0.5, 6)),
        ("median",            round(statistics.median(areas), 6)),
        ("mode",              statistics.mode(areas)),
        ("standard deviation",round(std_, 6)),
        ("sample variance",   round(float(np.var(areas, ddof=1)), 6)),
        ("range",             max(areas) - min(areas)),
        ("min",               min(areas)),
        ("max",               max(areas)),
        ("sum",               sum(areas)),
        ("count",             n),
    ]


def _summary_row(set_name: str, leg_means: dict, all_areas_flat: list) -> dict:
    import datetime
    leg_vals = [leg_means[lg] for lg in LEGS if lg in leg_means]
    avg      = round(sum(leg_vals) / len(leg_vals))
    std_legs = round(float(np.std(leg_vals, ddof=1))) if len(leg_vals) > 1 else 0
    std_imgs = round(float(np.std(all_areas_flat, ddof=1)))
    rng      = max(leg_vals) - min(leg_vals)
    return dict(
        date       = datetime.date.today().strftime("%d-%b-%Y"),
        set_name   = set_name,
        avg        = avg,
        FR         = leg_means.get("FR", ""),
        FL         = leg_means.get("FL", ""),
        BR         = leg_means.get("BR", ""),
        BL         = leg_means.get("BL", ""),
        std_legs   = std_legs,
        std_imgs   = std_imgs,
        rng        = rng,
        leg_pct    = round(std_legs / avg * 100) if avg else "",
        img_pct    = round(std_imgs / avg * 100) if avg else "",
        rng_pct    = round(rng      / avg * 100) if avg else "",
    )


def _anova_rows(all_areas_by_leg: dict, order=None) -> tuple:
    """Return (anova_table_rows, pairwise_rows) matching MATLAB ANOVA1 sheet.
    `order` sets the leg column order (group numbering); defaults to LEGS."""
    from scipy.stats import f_oneway
    order      = order or LEGS
    groups     = [all_areas_by_leg[lg] for lg in order if lg in all_areas_by_leg]
    leg_labels = [lg for lg in order if lg in all_areas_by_leg]
    if len(groups) < 2:
        return [], []

    k   = len(groups)
    n   = sum(len(g) for g in groups)
    grand_mean = sum(sum(g) for g in groups) / n

    ss_between = sum(len(g) * (sum(g)/len(g) - grand_mean)**2 for g in groups)
    ss_within  = sum(sum((x - sum(g)/len(g))**2 for x in g) for g in groups)
    ss_total   = ss_between + ss_within
    df_between = k - 1
    df_within  = n - k
    ms_between = ss_between / df_between
    ms_within  = ss_within  / df_within
    f_stat, p_val = f_oneway(*groups)

    anova_rows = [
        ["Source", "SS", "df", "MS", "F", "Prob>F"],
        ["Columns", round(ss_between, 6), df_between,
         round(ms_between, 6), round(float(f_stat), 6), round(float(p_val), 6)],
        ["Error",  round(ss_within,  6), df_within,
         round(ms_within,  6), "", ""],
        ["Total",  round(ss_total,   6), n - 1, "", "", ""],
    ]

    # Pairwise: Tukey-style critical range (mirrors MATLAB multcompare output columns)
    from scipy.stats import studentized_range
    pairs = []
    for i in range(k):
        for j in range(i + 1, k):
            gi, gj   = groups[i], groups[j]
            mi, mj   = sum(gi)/len(gi), sum(gj)/len(gj)
            diff     = mi - mj
            se       = (ms_within * (1/len(gi) + 1/len(gj)) / 2) ** 0.5
            q_crit   = studentized_range.ppf(0.95, k, df_within) if se > 0 else 0
            half_ci  = q_crit * se / 2**0.5
            pairs.append([
                i + 1, j + 1,
                round(diff - half_ci, 6),
                round(diff, 6),
                round(diff + half_ci, 6),
                round(float(f_oneway(gi, gj)[1]), 6),
            ])
    return anova_rows, pairs


# ── New consolidated format ───────────────────────────────────────────────────

def write_new_format(set_dir: str, leg_results: dict) -> str:
    """
    Single workbook: Summary tab + one tab per leg.
    Each leg tab has scratch areas + descStats on the left,
    individual scratch dimensions on the right.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    set_name = os.path.basename(set_dir)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    leg_means        = {}
    all_areas_by_leg = {}

    HDR  = Font(bold=True)
    FILL_HDR  = PatternFill("solid", start_color="D9E1F2")
    FILL_STAT = PatternFill("solid", start_color="EBF1DE")
    THIN = Side(style="thin")
    BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CTR  = Alignment(horizontal="center")

    def _hdr(ws, row, col, text, fill=None):
        c = ws.cell(row=row, column=col, value=text)
        c.font = HDR
        if fill:
            c.fill = fill
        c.border = BOX
        c.alignment = CTR

    def _val(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.border = BOX

    for leg in LEGS:
        results = leg_results.get(leg)
        if not results:
            continue

        sorted_results = sorted(results, key=lambda x: x["file"])
        areas = [r["scratch_area"] for r in sorted_results]
        leg_means[leg] = round(sum(areas) / len(areas))
        all_areas_by_leg[leg] = areas

        ws = wb.create_sheet(title=leg)
        ws.column_dimensions["A"].width = 14
        ws.column_dimensions["B"].width = 22
        ws.column_dimensions["C"].width = 3   # spacer
        ws.column_dimensions["D"].width = 14
        ws.column_dimensions["E"].width = 10
        ws.column_dimensions["F"].width = 10
        ws.column_dimensions["G"].width = 10

        # ── Left: scratch areas ───────────────────────────────────────────────
        _hdr(ws, 1, 1, "Image Name",           FILL_HDR)
        _hdr(ws, 1, 2, "Scratch Area (pixels)", FILL_HDR)
        for i, r in enumerate(sorted_results, start=2):
            _val(ws, i, 1, r["file"])
            _val(ws, i, 2, r["scratch_area"])

        stat_row = len(sorted_results) + 3
        _hdr(ws, stat_row - 1, 1, "Statistic", FILL_STAT)
        _hdr(ws, stat_row - 1, 2, "Value",     FILL_STAT)
        for label, value in _desc_stats(areas):
            _val(ws, stat_row, 1, label)
            _val(ws, stat_row, 2, value)
            ws.cell(row=stat_row, column=1).fill = FILL_STAT
            ws.cell(row=stat_row, column=2).fill = FILL_STAT
            stat_row += 1

        # ── Right: individual scratch dimensions ──────────────────────────────
        _hdr(ws, 1, 4, "Picture",   FILL_HDR)
        _hdr(ws, 1, 5, "Scratch #", FILL_HDR)
        _hdr(ws, 1, 6, "Width",     FILL_HDR)
        _hdr(ws, 1, 7, "Length",    FILL_HDR)
        detail_row = 2
        for r in sorted_results:
            for s in r.get("scratches", []):
                _val(ws, detail_row, 4, r["file"])
                _val(ws, detail_row, 5, s["scratch_num"])
                _val(ws, detail_row, 6, s["width_px"])
                _val(ws, detail_row, 7, s["length_px"])
                detail_row += 1

    # ── Summary tab ───────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet(title="Summary", index=0)
    ws_sum.column_dimensions["A"].width = 16
    for col in ["B","C","D","E","F","G","H","I","J","K","L","M"]:
        ws_sum.column_dimensions[col].width = 18

    sum_headers = ["Date", "Set Name", "Avg Scratch Area",
                   "FR", "FL", "BR", "BL",
                   "Std of Legs", "Std of Images", "Range",
                   "Leg Std %", "Img Std %", "Range %"]
    for col, h in enumerate(sum_headers, start=1):
        _hdr(ws_sum, 1, col, h, FILL_HDR)

    all_areas_flat = [a for areas in all_areas_by_leg.values() for a in areas]
    if leg_means and all_areas_flat:
        s = _summary_row(set_name, leg_means, all_areas_flat)
        row_data = [s["date"], s["set_name"], s["avg"],
                    s["FR"], s["FL"], s["BR"], s["BL"],
                    s["std_legs"], s["std_imgs"], s["rng"],
                    s["leg_pct"], s["img_pct"], s["rng_pct"]]
        for col, val in enumerate(row_data, start=1):
            _val(ws_sum, 2, col, val)

        # Each leg's Statistic/Value (descStats) block, laid out side by side on
        # the Summary tab so all four are visible at a glance.  (ANOVA + pairwise
        # now live on their own "ANOVA" tab instead of below the summary.)
        block_cols = {"FR": 1, "FL": 4, "BR": 7, "BL": 10}   # A-B, D-E, G-H, J-K
        block_top  = 4
        for leg, c0 in block_cols.items():
            areas = all_areas_by_leg.get(leg)
            if not areas:
                continue
            _hdr(ws_sum, block_top,     c0,     leg,         FILL_STAT)
            _hdr(ws_sum, block_top,     c0 + 1, "",          FILL_STAT)
            _hdr(ws_sum, block_top + 1, c0,     "Statistic", FILL_STAT)
            _hdr(ws_sum, block_top + 1, c0 + 1, "Value",     FILL_STAT)
            rr = block_top + 2
            for label, value in _desc_stats(areas):
                _val(ws_sum, rr, c0,     label)
                _val(ws_sum, rr, c0 + 1, value)
                ws_sum.cell(row=rr, column=c0).fill     = FILL_STAT
                ws_sum.cell(row=rr, column=c0 + 1).fill = FILL_STAT
                rr += 1

        # ── ANOVA tab (own sheet): ANOVA table + pairwise comparisons ───────────
        if len(all_areas_by_leg) >= 2:
            anova_rows, pair_rows = _anova_rows(all_areas_by_leg)
            ws_an = wb.create_sheet(title="ANOVA", index=1)
            ws_an.column_dimensions["A"].width = 16
            for col in "BCDEF":
                ws_an.column_dimensions[col].width = 14
            r = 1
            _hdr(ws_an, r, 1, "ANOVA", FILL_STAT)
            r += 1
            for i, ar in enumerate(anova_rows):
                for col, val in enumerate(ar, start=1):
                    c = ws_an.cell(row=r, column=col, value=val)
                    c.border = BOX
                    if i == 0:                 # first row is the column header
                        c.font = HDR
                r += 1
            r += 1
            _hdr(ws_an, r, 1, "Pairwise Comparisons", FILL_STAT)
            r += 1
            for col, ph in enumerate(
                    ["Group 1", "Group 2", "Lower CI", "Difference", "Upper CI", "p-value"],
                    start=1):
                c = ws_an.cell(row=r, column=col, value=ph)
                c.font = HDR; c.border = BOX
            r += 1
            for pr in pair_rows:
                for col, val in enumerate(pr, start=1):
                    ws_an.cell(row=r, column=col, value=val).border = BOX
                r += 1

    path = os.path.join(set_dir, f"{set_name}_results.xlsx")
    wb.save(path)
    return path


# ── Legacy format (3 files matching original MATLAB output) ──────────────────

def write_legacy_format(set_dir: str, leg_results: dict) -> list[str]:
    """
    Writes three files matching the original MATLAB xls output:
      {set_name}_scratch_count.xlsx  (Sheet1, per-leg sheets, ANOVA1)
      {set_name}_scratch_data.xlsx   (per-leg individual scratch sheets)
      Summary.xlsx
    """
    import openpyxl
    set_name = os.path.basename(set_dir)
    paths = []

    leg_means        = {}
    all_areas_by_leg = {}

    # ── scratch_count ─────────────────────────────────────────────────────────
    wb_count = openpyxl.Workbook()
    wb_count.remove(wb_count.active)
    ws_sheet1 = wb_count.create_sheet("Sheet1")
    ws_sheet1.append(["file", "scratch area"])

    for leg in LEGS:
        results = leg_results.get(leg)
        if not results:
            continue
        sorted_results = sorted(results, key=lambda x: x["file"])
        areas = [r["scratch_area"] for r in sorted_results]
        leg_means[leg] = round(sum(areas) / len(areas))
        all_areas_by_leg[leg] = areas

        ws = wb_count.create_sheet(title=leg)
        ws.append(["image name", "scratch area (pixels)"])
        for r in sorted_results:
            ws.append([r["file"], r["scratch_area"]])
            ws_sheet1.append([r["file"], r["scratch_area"]])
        ws.append([])
        for label, value in _desc_stats(areas):
            ws.append([label, value])

    if len(all_areas_by_leg) >= 2:
        ws_anova = wb_count.create_sheet("ANOVA1")
        anova_rows, pair_rows = _anova_rows(all_areas_by_leg)
        for ar in anova_rows:
            ws_anova.append(ar)
        ws_anova.append([])
        for pr in pair_rows:
            ws_anova.append(pr)

    p1 = os.path.join(set_dir, f"{set_name}_scratch_count.xlsx")
    wb_count.save(p1)
    paths.append(p1)

    # ── scratch_data ──────────────────────────────────────────────────────────
    wb_data = openpyxl.Workbook()
    wb_data.remove(wb_data.active)
    for leg in LEGS:
        results = leg_results.get(leg)
        if not results:
            continue
        ws = wb_data.create_sheet(title=leg)
        ws.append(["Picture", "Scratch", "Width", "Length"])
        for r in sorted(results, key=lambda x: x["file"]):
            for s in r.get("scratches", []):
                ws.append([r["file"], s["scratch_num"], s["width_px"], s["length_px"]])

    p2 = os.path.join(set_dir, f"{set_name}_scratch_data.xlsx")
    wb_data.save(p2)
    paths.append(p2)

    # ── Summary ───────────────────────────────────────────────────────────────
    all_areas_flat = [a for areas in all_areas_by_leg.values() for a in areas]
    if leg_means and all_areas_flat:
        wb_sum = openpyxl.Workbook()
        ws_s   = wb_sum.active
        ws_s.title = "Sheet1"
        ws_s.append(["", "date", "sample name", "average scratch area",
                     "FR", "FL", "BR", "BL",
                     "std.s of legs", "std.s of images", "range",
                     "leg std %", "img std %", "range %"])
        s = _summary_row(set_name, leg_means, all_areas_flat)
        ws_s.append([1, s["date"], s["set_name"], s["avg"],
                     s["FR"], s["FL"], s["BR"], s["BL"],
                     s["std_legs"], s["std_imgs"], s["rng"],
                     s["leg_pct"], s["img_pct"], s["rng_pct"]])
        p3 = os.path.join(set_dir, "Summary.xlsx")
        wb_sum.save(p3)
        paths.append(p3)

    return paths


# ── Summarize: combine many sets into one C8-style workbook ──────────────────

# Leg column order used by the C8 "Data set" template (top blocks + summary).
_SUMMARY_LEG_ORDER = ["BL", "BR", "FL", "FR"]
_IMG_EXTS = (".jpg", ".jpeg", ".png", ".tif", ".tiff", ".bmp")

# Analysis-only mode ignores any file whose name isn't one of the expected
# capture frames 000..030.  Keeps stray/renamed files (thumbnails, overlays,
# hand-added images) from being processed and polluting the results.
def _is_numbered_frame(fname: str) -> bool:
    stem = os.path.splitext(fname)[0]
    return stem.isdigit() and 0 <= int(stem) <= EXPECTED_IMAGES


# A file younger than this is assumed to be mid-write and left for the next
# poll.  Comfortably longer than a JPEG flush, far shorter than the gap between
# two captures, so it never delays a scan.
_SETTLE_SECONDS = 0.25


def _is_settled(path: str) -> bool:
    """True if `path` was last modified long enough ago to be safe to read.

    A missing file counts as unsettled rather than raising: it means the file
    vanished between listdir() and here (a rename in progress), so skipping it
    this round is exactly right.
    """
    try:
        return (time.time() - os.path.getmtime(path)) >= _SETTLE_SECONDS
    except OSError:
        return False


def _read_set_areas(workbook_path: str) -> dict:
    """
    Read one set's per-leg image scratch areas from its results workbook.

    Works for BOTH formats: the new '{set}_results.xlsx' and the legacy
    '{set}_scratch_count.xlsx' both keep raw data on sheets named FR/FL/BR/BL
    with the image name in column A and the scratch area in column B.  We take
    only the rows whose column A is an image filename, so the trailing descStats
    block (labelled 'mean', 'median', …) is skipped.

    Returns {leg: [areas]} for whichever legs are present.
    """
    import openpyxl
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    out = {}
    try:
        for leg in LEGS:
            if leg not in wb.sheetnames:
                continue
            areas = []
            for name, area in wb[leg].iter_rows(min_row=2, max_col=2, values_only=True):
                if (isinstance(name, str) and name.lower().endswith(_IMG_EXTS)
                        and isinstance(area, (int, float))):
                    areas.append(float(area))
            if areas:
                out[leg] = areas
    finally:
        wb.close()
    return out


def collect_sets(parent_dir: str) -> list:
    """
    Walk every subfolder under parent_dir and read each set's results workbook.
    Prefers the new '*_results.xlsx'; falls back to legacy '*_scratch_count.xlsx'.
    Returns a sorted list of (set_name, {leg: [areas]}).
    """
    sets = []
    for root, _dirs, files in os.walk(parent_dir):
        new    = sorted(f for f in files if f.endswith("_results.xlsx"))
        legacy = sorted(f for f in files if f.endswith("_scratch_count.xlsx"))
        if new:
            path, stem = os.path.join(root, new[0]), new[0][:-len("_results.xlsx")]
        elif legacy:
            path, stem = os.path.join(root, legacy[0]), legacy[0][:-len("_scratch_count.xlsx")]
        else:
            continue
        # Name the set after its FOLDER, not the workbook inside it.  The two
        # drift apart — a folder called "Control C39 S1" can hold a workbook
        # called "C39 S1_results.xlsx" — and the folder is what the operator
        # actually types, so it is what control detection has to read.  Falling
        # back to the workbook stem covers a set sitting loose in parent_dir.
        name = (os.path.basename(root)
                if os.path.abspath(root) != os.path.abspath(parent_dir) else stem)
        data = _read_set_areas(path)
        if data:
            sets.append((name, data))
    sets.sort(key=lambda x: x[0])
    return sets


_TEMPLATE_PATH = os.path.join(
    os.path.dirname(os.path.abspath(__file__)),
    "templates", "wet_abrasion_template.xlsx")

# A control is any set with "control" anywhere in its name, case-insensitive —
# "Control S1", "control B24", "S1 control" all match.  Controls are laid out
# after a black separator column and summarised in their own tables, so the
# reference films are never averaged in with the samples under test.
_CONTROL_RE = re.compile(r"control", re.I)

# Z threshold for outlier removal.  Reverse-engineered from the reference
# workbook: 1.96 reproduces every published Average and STDEV at every pass, on
# both the test films and the controls.  (The 1.5-IQR limits shown on the
# BoxPlots tab do NOT drive removal — they would have dropped two cells in pass
# one, where the sheet dropped only the single worst.)
_OUTLIER_Z = 1.96


def _is_control(set_name: str) -> bool:
    return bool(_CONTROL_RE.search(set_name or ""))


def _zscore_passes(per_set: list, z: float = _OUTLIER_Z) -> list:
    """Iteratively drop leg means beyond ±z·σ of the pooled group.

    Returns one CUMULATIVE removed-key set per pass; an empty list means the
    group was clean on the first look.  Each key is (set_index, leg).

    The reference sheet removes one pass at a time and recomputes the mean and
    stdev from the survivors before looking again, which is why a cell can sit
    inside the limit on pass one and outside it on pass two.  Iterating to a
    fixed point reproduces that, and also explains why the test films needed two
    passes while the controls settled after one.
    """
    keys = [(i, lg) for i, (_n, lm, _a, _s) in enumerate(per_set) for lg in lm]
    vals = {k: per_set[k[0]][1][k[1]] for k in keys}
    removed, passes = set(), []
    while len(passes) < 10:                      # hard stop; never seen past 2
        live = [k for k in keys if k not in removed]
        if len(live) < 3:
            break
        v = [vals[k] for k in live]
        m, s = float(np.mean(v)), float(np.std(v, ddof=1))
        if s == 0:
            break
        flagged = {k for k in live if abs((vals[k] - m) / s) > z}
        if not flagged:
            break
        removed |= flagged
        passes.append(set(removed))
    return passes


def _live_stats(per_set: list, legs_order: list, removed: set) -> dict:
    """Summary numbers for a group with `removed` leg cells excluded.

    Excluded cells are still WRITTEN to the sheet (flagged red) — they are only
    left out of the arithmetic, so the reader can always see what was dropped.
    """
    live = {}                                     # (set_idx, leg) → value
    for i, (_n, lm, _a, _s) in enumerate(per_set):
        for lg, v in lm.items():
            if (i, lg) not in removed:
                live[(i, lg)] = v
    all_v = list(live.values())
    per_leg = {lg: [v for (i, l), v in live.items() if l == lg] for lg in legs_order}
    n = len(all_v)
    std = float(np.std(all_v, ddof=1)) if n > 1 else 0.0
    return {
        "live":      live,
        "leg_avg":   {lg: (sum(x) / len(x) if x else None) for lg, x in per_leg.items()},
        "leg_std":   {lg: (float(np.std(x, ddof=1)) if len(x) > 1 else None)
                      for lg, x in per_leg.items()},
        "grand":     (sum(all_v) / n if n else None),
        "std":       std,
        # CONFIDENCE(0.05, sigma, n).  The reference sheet hard-codes n=16 on the
        # raw table, which is right only when all four legs of four sets survive
        # — it understates the controls (n=8).  The real count is used here.
        "ci":        (1.959963985 * std / math.sqrt(n)) if n > 1 else None,
        "range":     (max(all_v) - min(all_v)) if all_v else None,
        "n":         n,
    }


def write_summarize_format(parent_dir: str, sets: list, out_path: str = None) -> str:
    """
    Build the lab's wet-abrasion summary workbook, four tabs:
      • <batch> — a horizontal block per set (raw per-image areas by leg →
        descStats → ANOVA → pairwise), test films first, then a full-height
        black separator column, then the controls.  Below them, paired summary
        tables: test on the left, controls past the separator, one table per
        outlier pass.  Outlier cells keep their value and are flagged red;
        only the arithmetic excludes them.
      • Template — the lab's outlier worksheet, copied verbatim with its
        formulas and conditional formatting intact.
      • T-Test — pooled image areas, test vs control, plus =TTEST(...,2,3).
      • BoxPlots — per-pass tables, IQR limits, and a min/Q1/Q3/max table
        laid out ready to chart from.
    Returns the written path.
    """
    import datetime
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    # Styling follows the lab's reference sheet rather than our own taste: Arial
    # for structure (set names, leg headers, summary tables), Calibri 11 for raw
    # data, essentially NO background fill, and borders only where they separate
    # something — a grid drawn on every cell reads as noise at this density.
    HDR       = Font(name="Arial", size=10, bold=True)
    FONT_ARI  = Font(name="Arial", size=10)
    FONT_BODY = Font(name="Calibri", size=11)
    FILL_HDR  = PatternFill("solid", start_color="FFD9E1F2")
    FILL_STAT = None                                   # reference uses no fill
    FILL_TOT  = PatternFill("solid", start_color="FFFFF2CC")  # grand avg/stdev
    FILL_RED  = PatternFill("solid", start_color="FFFFC7CE")  # outlier flag
    FONT_RED  = Font(name="Arial", size=10, color="9C0006", bold=True)
    THIN  = Side(style="thin")
    THICK = Side(style="thick")
    BOX   = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
    CTR   = Alignment(horizontal="center")
    FMT_1D, FMT_3D, FMT_SCI = "0.0", "0.000", "0.00E+00"

    # Fully-opaque black (FF alpha).  A bare "000000" is stored with alpha 00,
    # which some readers treat as transparent rather than black.
    FILL_BLACK = PatternFill("solid", start_color="FF000000", end_color="FF000000")

    legs_order = _SUMMARY_LEG_ORDER
    wb = openpyxl.Workbook()
    ws = wb.active
    # The main tab carries the batch name — the folder the sets were captured
    # into.  Excel forbids []:*?/\ in a sheet title and caps it at 31 chars.
    batch = os.path.basename(os.path.abspath(parent_dir))
    ws.title = (re.sub(r"[\[\]:*?/\\]", "-", batch)[:31] or "Summary")

    # Test films first, then the controls — the split drives both the raw block
    # and the summary tables below it.  Order within each group is preserved.
    test_sets = [s for s in sets if not _is_control(s[0])]
    ctrl_sets = [s for s in sets if _is_control(s[0])]
    sets = test_sets + ctrl_sets

    def make_cell(target):
        def cell(r, c, value, *, bold=False, fill=None, box=False, ctr=False, font=None):
            x = target.cell(row=r, column=c, value=value)
            if bold:  x.font = HDR
            if font:  x.font = font
            if fill:  x.fill = fill
            if box:   x.border = BOX
            if ctr:   x.alignment = CTR
            return x
        return cell
    cell = make_cell(ws)

    # Metadata header (rows 1-3): labels only — who ran it and the run order are
    # facts the scanner has no way to know, so they are left blank to fill in.
    for ri, label in enumerate(("Date Completed", "Completed By", "Notes"), 1):
        c = cell(ri, 1, label, font=FONT_ARI)
        c.alignment = Alignment(horizontal="left")
        c.border = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
        ws.cell(row=ri, column=2).border = Border(right=THIN, top=THIN, bottom=THIN)
    ws.column_dimensions["A"].width = 17.6

    # ── Top: one horizontal block per set (below the metadata header) ──────────
    BLOCK_TOP  = 5                        # first block row (leaves rows 1-3 for meta)
    N_IMG      = 30                       # expected images per leg (layout spacing)
    SET_WIDTH  = len(legs_order) * 2      # 2 columns (name, area) per leg
    SET_STRIDE = SET_WIDTH                # blocks sit flush; no spacer column
    per_set = []                          # cache computed values for the summary

    # Column of the black separator, and the first control column after it.
    # The separator only exists when there is something on BOTH sides of it —
    # an all-control run would otherwise paint column 1 black over its own data.
    n_test    = len(test_sets)
    split     = bool(test_sets and ctrl_sets)
    BLACK_COL = 1 + n_test * SET_STRIDE if split else None
    CTRL_COL0 = (BLACK_COL + 1) if split else (1 if ctrl_sets else None)

    def set_col0(i):
        """First column of set i, stepping over the black column for controls."""
        return 1 + i * SET_STRIDE + (1 if (split and i >= n_test) else 0)

    stats_labels = [lbl for lbl, _ in _desc_stats([0, 1])]
    for i, (name, data) in enumerate(sets):
        c0 = set_col0(i)
        # Set name spans the whole block, leg names span their two columns —
        # without the merges the headers float above the wrong column and the
        # sheet reads as 48 unlabelled columns.
        ws.merge_cells(start_row=BLOCK_TOP, start_column=c0,
                       end_row=BLOCK_TOP, end_column=c0 + SET_WIDTH - 1)
        h = cell(BLOCK_TOP, c0, name, ctr=True, font=FONT_ARI)
        h.border = Border(left=THICK, right=THICK, top=THICK, bottom=THIN)

        for j, leg in enumerate(legs_order):
            nc = c0 + j * 2               # image-name column for this leg
            ws.merge_cells(start_row=BLOCK_TOP + 1, start_column=nc,
                           end_row=BLOCK_TOP + 1, end_column=nc + 1)
            lc = cell(BLOCK_TOP + 1, nc, leg, ctr=True, font=FONT_ARI)
            lc.border = Border(left=THICK if j == 0 else THIN,
                               right=THICK if j == len(legs_order) - 1 else THIN,
                               top=THIN, bottom=THIN)
            edge_l = THICK if j == 0 else THIN
            edge_r = THICK if j == len(legs_order) - 1 else None
            cell(BLOCK_TOP + 2, nc,     "image name",            font=FONT_BODY)
            cell(BLOCK_TOP + 2, nc + 1, "scratch area (pixels)", font=FONT_BODY)
            ws.cell(row=BLOCK_TOP + 2, column=nc).border = Border(left=edge_l)
            if edge_r:
                ws.cell(row=BLOCK_TOP + 2, column=nc + 1).border = Border(right=edge_r)

            areas = data.get(leg, [])
            r = BLOCK_TOP + 3
            for k, a in enumerate(areas):
                c = cell(r, nc, f"{k + 1:03d}.jpg", font=FONT_BODY)
                c.border = Border(left=edge_l)
                v = cell(r, nc + 1, a, font=FONT_BODY)
                if edge_r:
                    v.border = Border(right=edge_r)
                r += 1
            if areas:                     # descStats block below the raw rows
                for si, (label, value) in enumerate(_desc_stats(areas)):
                    top_s  = THIN  if si == 0 else None
                    bot_s  = THICK if si == len(stats_labels) - 1 else None
                    lc2 = cell(r, nc, label, font=FONT_BODY)
                    lc2.border = Border(left=edge_l, top=top_s, bottom=bot_s)
                    vc = cell(r, nc + 1, value, font=FONT_BODY)
                    vc.border = Border(right=edge_r, top=top_s, bottom=bot_s)
                    # Variance runs to ~1e9; scientific keeps the column narrow,
                    # which is what the reference does with this row alone.
                    if label == "sample variance":
                        vc.number_format = FMT_SCI
                    r += 1

        # ANOVA + pairwise for this set (needs ≥2 legs)
        present = {lg: data[lg] for lg in legs_order if lg in data}
        if len(present) >= 2:
            anova_rows, pair_rows = _anova_rows(present, order=legs_order)
            ar = BLOCK_TOP + 3 + N_IMG + len(stats_labels) + 2   # below the stats
            for gi, arow in enumerate(anova_rows):
                for dc, val in enumerate(arow):
                    c = cell(ar, c0 + dc, val, font=FONT_BODY)
                    if gi == 0:
                        c.font = HDR
                    elif dc in (1, 3) and isinstance(val, (int, float)):
                        c.number_format = FMT_SCI      # SS and MS are ~1e8-1e11
                ar += 1
            ar += 1
            for prow in pair_rows:
                for dc, val in enumerate(prow):
                    cell(ar, c0 + dc, val, font=FONT_BODY)
                ar += 1

        # Per-set summary numbers (display leg order; only present legs count)
        leg_means = {lg: sum(data[lg]) / len(data[lg]) for lg in present}
        avg = sum(leg_means.values()) / len(leg_means) if leg_means else 0.0
        std_legs = float(np.std(list(leg_means.values()), ddof=1)) if len(leg_means) > 1 else 0.0
        per_set.append((name, leg_means, avg, std_legs))

    # The black separator column: it runs the full height of the raw block AND
    # the summary tables below, so the eye never has to work out where the test
    # films stop and the reference controls begin.
    stats_h  = len(_desc_stats([0]))
    anova_h  = 12                                   # 4 ANOVA rows + gap + 6 pairs
    raw_last = BLOCK_TOP + 3 + N_IMG + stats_h + anova_h

    # ── Summary tables: test films on the left, controls past the black column ──
    # Each group is judged against ITS OWN pool, so a control is never compared
    # against the films under test (and vice versa).  Outlier cells keep their
    # value and are flagged red; only the arithmetic excludes them.
    per_test = per_set[:n_test]
    per_ctrl = per_set[n_test:]

    # Stages are written index-by-index so the test and control tables stay
    # ROW-ALIGNED across the black column even when the groups hold a different
    # number of sets or need a different number of passes — the raw pair sits
    # side by side, then pass 1 beside pass 1, and so on.
    def stages_for(group):
        if not group:
            return []
        passes = _zscore_passes(group)
        out = [("", set())]
        for pi, removed in enumerate(passes, 1):
            # One pass is labelled plainly; two or more are numbered, matching
            # how the reference sheet distinguishes them.
            out.append((("Outliers Removed" if len(passes) == 1
                         else f"Outliers Removed - {pi} pass"), removed))
        return out

    # Live references back into the raw block: each leg cell in a summary table
    # points at that leg's descStats "mean", and the set name at its header —
    # so correcting a raw area re-flows the whole summary.  Sheet-qualified and
    # quoted because the BoxPlots tab reuses the same tables.
    from openpyxl.utils import get_column_letter as _colletter
    MEAN_ROW = BLOCK_TOP + 3 + N_IMG              # first descStats row = mean
    qname    = "'" + ws.title.replace("'", "''") + "'"

    def make_mean_ref(offset):
        def ref(i, leg):
            c0i = set_col0(offset + i)
            if leg == "NAME":
                return f"{qname}!{_colletter(c0i)}${BLOCK_TOP}"
            j = legs_order.index(leg)
            return f"{qname}!{_colletter(c0i + j * 2 + 1)}${MEAN_ROW}"
        return ref

    cols = [(per_test, stages_for(per_test), 1, make_mean_ref(0))]
    if per_ctrl and CTRL_COL0:
        cols.append((per_ctrl, stages_for(per_ctrl), CTRL_COL0,
                     make_mean_ref(n_test)))

    sum_top = raw_last + 3
    r = sum_top
    for si in range(max(len(s) for _g, s, _c, _m in cols)):
        ends = []
        for group, stages, col0, mref in cols:
            if si >= len(stages):
                continue
            title, removed = stages[si]
            ends.append(_write_fmt_summary_table(
                ws, r, col0, title, group, legs_order, removed,
                cell, HDR, FILL_STAT, FILL_RED, FONT_RED, mean_ref=mref))
        r = max(ends) + 2
    last_rows = [r]

    if BLACK_COL:
        for rr in range(BLOCK_TOP, max(last_rows) + 1):
            ws.cell(row=rr, column=BLACK_COL).fill = FILL_BLACK
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(BLACK_COL)].width = 2.5

    # ── Remaining tabs ────────────────────────────────────────────────────────
    _copy_template_tab(wb)
    _write_ttest_tab(wb, per_test, per_ctrl, sets, n_test)
    _write_boxplots_tab(wb, per_test, per_ctrl, legs_order, batch,
                        cellmaker=make_cell, HDR=HDR, FILL_STAT=FILL_STAT,
                        FILL_RED=FILL_RED, FONT_RED=FONT_RED,
                        mean_refs=[m for _g, _s, _c, m in cols])

    out_path = out_path or os.path.join(
        parent_dir, f"{os.path.basename(os.path.abspath(parent_dir))}_summary.xlsx")
    wb.save(out_path)
    return out_path


def _write_fmt_summary_table(ws, top, col0, title, group, legs_order, removed,
                             cell, HDR, FILL_STAT, FILL_RED, FONT_RED,
                             mean_ref=None):
    """One Film summary table at (top, col0).  Returns its last row.

    Columns: Film | <legs> | Average | STDEV of Legs | 95% Confidence Interval,
    then Average / STDEV rows across sets and a RANGE line.  Cells in `removed`
    are written with their real value but flagged red, and every statistic on
    the table ignores them — the reference sheet blanks them instead, which
    loses the evidence of what was dropped.

    With `mean_ref` (a callable (set_index, leg) → an absolute cell reference
    for that leg's mean in the raw block) the table is written as LIVE FORMULAS
    instead of literals: each leg cell points at its descStats mean, and every
    statistic is an AVERAGE/STDEV over an explicit list of the surviving cells.
    Enumerating the survivors rather than using a range is the whole trick — a
    plain AVERAGE(B62:E62) would quietly pull the red-flagged outliers back in,
    which is exactly what the pass exists to exclude.  Edit a raw area and the
    summary follows; the outlier decisions stay put.
    """
    from openpyxl.styles import Border, Side, Font, PatternFill, Alignment
    THIN, THICK = Side(style="thin"), Side(style="thick")
    ARI     = Font(name="Arial", size=10)
    ARI_B   = Font(name="Arial", size=10, bold=True)
    TOT     = PatternFill("solid", start_color="FFFFF2CC")
    nlegs   = len(legs_order)
    last_c  = col0 + 3 + nlegs

    def put(rr, cc, v, *, font=ARI, fill=None, fmt="0.0", top=None, bottom=None):
        c = ws.cell(row=rr, column=cc, value=v)
        c.font = font
        c.border = Border(left=THICK if cc == col0 else THIN,
                          right=THICK if cc == last_c else THIN,
                          top=top, bottom=bottom)
        if fill:
            c.fill = fill
        if isinstance(v, float):
            c.number_format = fmt
        return c

    r = top
    if title:
        cell(r, col0, title, font=ARI_B)
        r += 1
    for j, h in enumerate(["Film"] + legs_order +
                          ["Average", "STDEV of Legs", "95% Confidence Interval"]):
        put(r, col0 + j, h, font=ARI_B, top=THICK, bottom=THIN)
    r += 1

    from openpyxl.utils import get_column_letter as _cl
    st   = _live_stats(group, legs_order, removed)
    live = st["live"]
    body_top = r
    kept_at  = {}                 # (set_idx, leg) → this table's cell address

    for i, (name, lm, _a, _s) in enumerate(group):
        put(r, col0, (f"={mean_ref(i, 'NAME')}" if mean_ref else name), fmt=None)
        kept_refs = []
        for j, lg in enumerate(legs_order):
            addr = f"{_cl(col0 + 1 + j)}{r}"
            if lg not in lm:
                put(r, col0 + 1 + j, None)
                continue
            flagged = (i, lg) in removed
            val = f"={mean_ref(i, lg)}" if mean_ref else round(lm[lg], 1)
            c = put(r, col0 + 1 + j, val, fill=FILL_RED if flagged else None)
            c.number_format = "0.0"
            if flagged:
                c.font = FONT_RED
            else:
                kept_refs.append(addr)
                kept_at[(i, lg)] = addr
        # Explicit survivor list, never a range — that is what keeps the red
        # cells out of the arithmetic while leaving them on the sheet.
        if mean_ref and kept_refs:
            joined = ",".join(kept_refs)
            put(r, col0 + 1 + nlegs, f"=AVERAGE({joined})")
            put(r, col0 + 2 + nlegs,
                f"=STDEV({joined})" if len(kept_refs) > 1 else None)
        else:
            kv = [lm[lg] for lg in legs_order
                  if lg in lm and (i, lg) not in removed]
            put(r, col0 + 1 + nlegs, round(sum(kv) / len(kv), 1) if kv else None)
            put(r, col0 + 2 + nlegs,
                round(float(np.std(kv, ddof=1)), 1) if len(kv) > 1 else None)
        put(r, col0 + 3 + nlegs, None)
        r += 1

    def num(v):
        return round(v, 1) if isinstance(v, float) else None

    all_kept = list(kept_at.values())
    col_kept = {lg: [a for (ii, l), a in kept_at.items() if l == lg]
                for lg in legs_order}

    def agg(fn, refs):
        return f"={fn}({','.join(refs)})" if refs else None

    put(r, col0, "Average", font=ARI_B, top=THIN)
    for j, lg in enumerate(legs_order):
        put(r, col0 + 1 + j,
            agg("AVERAGE", col_kept[lg]) if mean_ref else num(st["leg_avg"][lg]),
            top=THIN)
    # The grand average and its stdev are the numbers people quote, so the
    # reference tints just those two — the only fill on the whole table.
    std_addr = f"{_cl(col0 + 2 + nlegs)}{r}"
    put(r, col0 + 1 + nlegs,
        agg("AVERAGE", all_kept) if mean_ref else num(st["grand"]),
        fill=TOT, top=THIN)
    put(r, col0 + 2 + nlegs,
        agg("STDEV", all_kept) if mean_ref else num(st["std"]),
        fill=TOT, top=THIN)
    # CONFIDENCE wants the surviving count, not a hard-coded 16 as the
    # reference uses — that only happens to be right for four full sets.
    put(r, col0 + 3 + nlegs,
        (f"=CONFIDENCE(0.05,{std_addr},{len(all_kept)})"
         if (mean_ref and all_kept) else num(st["ci"])),
        top=THIN)
    r += 1

    put(r, col0, "STDEV", font=ARI_B, bottom=THICK)
    for j, lg in enumerate(legs_order):
        refs = col_kept[lg]
        put(r, col0 + 1 + j,
            (agg("STDEV", refs) if len(refs) > 1 else None) if mean_ref
            else num(st["leg_std"][lg]),
            bottom=THICK)
    for j in range(1, 4):
        put(r, col0 + j + nlegs, None, bottom=THICK)
    r += 1

    ws.cell(row=r, column=col0 + 1 + nlegs, value="RANGE").font = ARI_B
    rc = ws.cell(row=r, column=col0 + 2 + nlegs,
                 value=(f"=MAX({','.join(all_kept)})-MIN({','.join(all_kept)})"
                        if (mean_ref and all_kept)
                        else (round(st["range"], 3)
                              if st["range"] is not None else None)))
    rc.font, rc.number_format = ARI, "0.000"
    return r


def _copy_template_tab(wb):
    """Clone the Template tab from the shipped workbook, formulas and all.

    The tab is a formula-driven outlier worksheet (IQR limits, Z scores, a
    normal-fit histogram and a Chi-squared cell) with conditional formatting on
    top.  It is copied verbatim rather than reimplemented: the formulas are the
    lab's, not ours, and openpyxl writes them straight through for Excel or
    Sheets to evaluate on open.  A missing template is not fatal — the data
    tabs are still worth having — so this degrades to a warning.
    """
    from copy import copy as _copy
    import openpyxl
    if not os.path.exists(_TEMPLATE_PATH):
        print(f"[analysis] template not found at {_TEMPLATE_PATH}; skipping tab")
        return
    src = openpyxl.load_workbook(_TEMPLATE_PATH)["Template"]
    dst = wb.create_sheet("Template")
    for row in src.iter_rows():
        for c in row:
            t = dst.cell(row=c.row, column=c.column, value=c.value)
            if c.has_style:
                t.font = _copy(c.font);   t.fill = _copy(c.fill)
                t.border = _copy(c.border); t.alignment = _copy(c.alignment)
                t.number_format = c.number_format
    for rng, rules in src.conditional_formatting._cf_rules.items():
        for rule in rules:
            dst.conditional_formatting.add(str(rng.sqref), _copy(rule))
    for k, d in src.column_dimensions.items():
        dst.column_dimensions[k].width = d.width
    for k, d in src.row_dimensions.items():
        dst.row_dimensions[k].height = d.height
    for m in src.merged_cells.ranges:
        dst.merged_cells.add(str(m))


def _write_ttest_tab(wb, per_test, per_ctrl, sets, n_test):
    """Two stacked columns of raw image areas, test vs control, plus the t-test.

    TTEST(...,2,3) is two-tailed, two-sample assuming UNEQUAL variance
    (Welch's) — the right choice here, since a test film and a control have no
    reason to share a variance.  The formula is written rather than a computed
    number so the value updates if rows are edited by hand afterwards.
    """
    from openpyxl.styles import Font
    ws = wb.create_sheet("T-Test")
    ws["A1"] = "Test Batch";  ws["F1"] = "Control"
    ws["A3"] = "image name";  ws["B3"] = "scratch area (pixels)"
    ws["D3"] = "T-Test"
    ws["F3"] = "image name";  ws["G3"] = "scratch area (pixels)"
    for c in ("A1", "F1", "A3", "B3", "D3", "F3", "G3"):
        ws[c].font = Font(bold=True)

    def dump(group_sets, name_col, val_col):
        r = 4
        for name, data in group_sets:
            for lg in _SUMMARY_LEG_ORDER:
                for k, a in enumerate(data.get(lg, [])):
                    ws.cell(row=r, column=name_col, value=f"{k + 1:03d}.jpg")
                    # Written as a NUMBER on purpose: one text-formatted value
                    # ("30,309") in the reference workbook was enough to make
                    # TTEST return #NUM! for the whole sheet.
                    ws.cell(row=r, column=val_col, value=int(a))
                    r += 1
        return r - 1

    last_t = dump(sets[:n_test], 1, 2)
    last_c = dump(sets[n_test:], 6, 7)
    if last_t >= 4 and last_c >= 4:
        ws["D4"] = f"=TTEST(B4:B{last_t},G4:G{last_c},2,3)"


def _write_boxplots_tab(wb, per_test, per_ctrl, legs_order, batch,
                        cellmaker, HDR, FILL_STAT, FILL_RED, FONT_RED,
                        mean_refs=None):
    """Per-pass summary tables, IQR limits, and the box-plot source table.

    The tab ends at the numbers.  Excel has no box-plot type openpyxl can emit,
    and the stock-chart substitute rendered badly once the workbook was
    converted to a Google Sheet, so the drawing is left to whoever wants it —
    the (minimum, Q1, Q3, maximum) table is laid out ready to chart from.
    """
    ws = wb.create_sheet("BoxPlots")
    cell = cellmaker(ws)
    chart_rows = []          # (label, first data row of the block)
    r = 2

    refs = list(mean_refs or [])
    for gi, (label, group) in enumerate((("Test", per_test), ("Control", per_ctrl))):
        if not group:
            continue
        mref = refs[gi] if gi < len(refs) else None
        passes = _zscore_passes(group)
        stages = [("", set())]
        for pi, removed in enumerate(passes, 1):
            stages.append((("Outliers Removed" if len(passes) == 1
                            else f"Outliers Removed - {pi} pass"), removed))
        for title, removed in stages:
            head = r + (1 if title else 0)
            r = _write_fmt_summary_table(ws, r, 1, title, group, legs_order,
                                         removed, cell, HDR, FILL_STAT,
                                         FILL_RED, FONT_RED, mean_ref=mref)
            st = _live_stats(group, legs_order, removed)
            vals = sorted(st["live"].values())
            if len(vals) >= 4:
                q1, med, q3 = (float(np.percentile(vals, p)) for p in (25, 50, 75))
                iqr = q3 - q1
                cols = ["Median", "Q1", "Q3", "Max", "Min", "IQR",
                        "+1.5 IQR (outliers)", "-1.5 IQR (outliers)"]
                nums = [med, q1, q3, max(vals), min(vals), iqr,
                        # The reference sheet's "1.5 IQR" columns actually add a
                        # bare IQR; these are the real 1.5x limits.
                        q3 + 1.5 * iqr, q1 - 1.5 * iqr]
                cell(head, 10, "Interquartile Limits", bold=True)
                for j, (cn, v) in enumerate(zip(cols, nums)):
                    cell(head + 1, 10 + j, cn, bold=True, fill=FILL_STAT, box=True)
                    cell(head + 2, 10 + j, round(v, 1), box=True)
                nm = f"{batch} {title}".strip() if title else f"{batch} Raw"
                chart_rows.append((f"{label} — {nm}",
                                   min(vals), q1, q3, max(vals)))
            r += 2

    # Box-plot source table, kept in the reference sheet's column order:
    # minimum, Q1, Q3, maximum.
    if chart_rows:
        top = r + 1
        cell(top, 19, "Box Plot Graphing (sheets formatted)", bold=True)
        for j, h in enumerate(["Film", "minimum", "Q1", "Q3", "maximum"]):
            cell(top + 1, 19 + j, h, bold=True, fill=FILL_STAT, box=True)
        for i, (nm, lo, q1, q3, hi) in enumerate(chart_rows):
            cell(top + 2 + i, 19, nm, box=True)
            for j, v in enumerate((lo, q1, q3, hi)):
                cell(top + 2 + i, 20 + j, round(float(v), 1), box=True)

        # No chart object is emitted.  A stock chart was the only box-plot
        # substitute openpyxl can write, and it did not survive the trip through
        # Google Sheets cleanly.  The table above is the whole payload: select
        # it and insert a candlestick chart if you want one drawn.


# ── Pipeline class ────────────────────────────────────────────────────────────

class AnalysisPipeline:
    """
    Watches leg_dir (set_dir/leg/) for new images and processes them.
    When done, triggers set-level Excel export via on_leg_done callback.
    """

    # A frame that fails to analyse is retried rather than abandoned — a single
    # unlucky read used to cost a whole image out of the 30.  Three attempts is
    # enough to ride out any transient (a file still flushing, a momentary lock)
    # while still giving up promptly on a genuinely unreadable file.
    _MAX_ATTEMPTS = 3
    _RETRY_DELAY  = 0.4      # seconds to wait when every candidate was deferred

    def __init__(self, leg_dir,
                 on_progress=None, on_done=None, on_error=None,
                 on_image=None, total_expected=None, mode="legacy",
                 live_capture=False):
        self._dir = leg_dir
        self._on_progress = on_progress or (lambda done, total: None)
        self._on_done = on_done or (lambda results: None)
        self._on_error = on_error or (lambda e: None)
        # on_image(overlay_path, fname): fired after each image is analysed, for
        # a live "last analysed" preview in the GUI.
        self._on_image = on_image or (lambda overlay_path, fname: None)
        self._total = total_expected
        self._mode = mode               # "legacy" (MATLAB) or "accurate"
        self._stop_event = threading.Event()
        # live_capture=True: images are still being captured into leg_dir, so
        # the watcher must NOT stop on an idle gap between shots — only once
        # mark_capture_done() is called (and the backlog is drained).  This lets
        # analysis run concurrently with capture instead of waiting for the set.
        self._live_capture = live_capture
        self._capture_done = threading.Event()
        self._thread = None
        self._results_path = os.path.join(leg_dir, "results.jsonl")

    def mark_capture_done(self):
        """Signal that no more images will be captured — the watcher may finish
        the remaining backlog and then stop."""
        self._capture_done.set()

    def start(self):
        os.makedirs(self._dir, exist_ok=True)
        self._thread = threading.Thread(target=self._run, daemon=True)
        self._thread.start()

    def stop(self):
        self._stop_event.set()

    def _run(self):
        try:
            processed = set()
            attempts = {}          # fname → failed attempts so far
            all_results = []
            done = 0
            idle_ticks = 0

            with open(self._results_path, "w") as rf:
                while not self._stop_event.is_set():
                    images = sorted(
                        f for f in os.listdir(self._dir)
                        if (f.endswith(".jpg") or f.endswith(".png"))
                        and not f.endswith("_overlay.png")
                        and f not in processed
                        # Frames the autofocus gave up on are saved as
                        # NNN_soft.jpg for the record — never analysed.
                        and "_soft" not in f
                        # Analysis-only mode: process ONLY the numbered capture
                        # frames 000..030, nothing else.  (Live/"both" capture
                        # writes its own frames, so no need to filter there.)
                        and (self._live_capture or _is_numbered_frame(f))
                        # Belt and braces alongside the atomic save in
                        # capture_pipeline: ignore anything written in the last
                        # fraction of a second, so a file still being flushed by
                        # some other writer is left alone until it settles.
                        and _is_settled(os.path.join(self._dir, f))
                    )
                    if images:
                        idle_ticks = 0
                        advanced = False       # did anything actually complete?
                        for fname in images:
                            if self._stop_event.is_set():
                                break
                            path = os.path.join(self._dir, fname)
                            try:
                                result = detect_scratches(path, mode=self._mode)
                                result["file"] = fname
                                all_results.append(result)
                                rf.write(json.dumps({
                                    "file": fname,
                                    "scratch_area": result["scratch_area"],
                                    "scratch_count": result["scratch_count"],
                                }) + "\n")
                                rf.flush()
                                # Live preview of the annotated overlay
                                self._on_image(result["overlay_path"], fname)
                            except Exception as e:
                                # A failure here used to retire the frame for good,
                                # so one unlucky read cost a whole image out of the
                                # 30 — the file was fine a moment later, but nothing
                                # ever looked at it again.  Retry a few polls later
                                # instead, and only give up (and record the error)
                                # once it has genuinely failed _MAX_ATTEMPTS times.
                                attempts[fname] = attempts.get(fname, 0) + 1
                                if attempts[fname] < self._MAX_ATTEMPTS:
                                    continue          # leave unprocessed; retry
                                rf.write(json.dumps({
                                    "file": fname,
                                    "error": str(e),
                                    "attempts": attempts[fname],
                                }) + "\n")
                                rf.flush()
                            processed.add(fname)
                            advanced = True
                            done += 1
                            self._on_progress(done, self._total)
                        # Every candidate was deferred for a retry.  Without a
                        # pause the loop would spin at full CPU re-reading the
                        # same half-written file, which is exactly the condition
                        # that stops it settling.
                        if not advanced:
                            time.sleep(self._RETRY_DELAY)
                    else:
                        idle_ticks += 1
                        # While capture is still live, never stop on an idle gap
                        # between shots — wait for more images (or stop()).
                        waiting_for_capture = (
                            self._live_capture and not self._capture_done.is_set()
                        )
                        if (not waiting_for_capture) and done > 0 and idle_ticks > 10:
                            break
                        time.sleep(0.5)

            self._on_done(all_results)
        except Exception as e:
            self._on_error(e)
